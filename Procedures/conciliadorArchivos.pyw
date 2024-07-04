import tkinter as tk
from tkinter import filedialog, ttk
from ttkthemes import ThemedTk
import sys
import os
from os import scandir
from os.path import abspath, join, relpath
from openpyxl import Workbook
import argparse

# pip install ttkthemes
# pip install openpyxl



class FileReconcilerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Conciliador de archivos")
        self.root.resizable(False, False)  # Deshabilitar el redimensionamiento
        self.root.geometry("800x600")  # Establecer el tamaño de la ventana principal

        def get_resource_path(relative_path):
            """ Obtiene la ruta del archivo de recursos, considerando si se está ejecutando desde un ejecutable empaquetado. """
            if hasattr(sys, '_MEIPASS'):
                return os.path.join(sys._MEIPASS, relative_path)
            return os.path.join(os.path.abspath("."), relative_path)
        # # Obtener la ruta del archivo que se está ejecutando
        # ruta_archivo_actual = os.path.abspath(__file__)

        # print("ruta_archivo_actual:", ruta_archivo_actual)

        def get_executable_directory():
            # Obtiene la ubicación del directorio donde se encuentra el ejecutable
            if getattr(sys, 'frozen', False):
                # Modo ejecutable (compilado)
                return os.path.dirname(sys.executable)
            else:
                # Modo script (desarrollo)
                return os.path.dirname(os.path.abspath(__file__))


        def toggle_mode():
            if mode_switch.instate(["selected"]):  # Cuando esté en el estado "selected"
                style.theme_use("forest-light")
                root.configure(bg=light_bg_color)
            else:
                style.theme_use("forest-dark")
                root.configure(bg=dark_bg_color)
            


        # Establecer dark mode
        style = ttk.Style(root)

        theme_dir = os.path.join(get_executable_directory(), "Assets/Themes")
        theme_files = ["forest-light.tcl", "forest-dark.tcl"]

        for theme_file in theme_files:
            theme_path = os.path.join(theme_dir, theme_file)
            root.tk.call("source", theme_path)

        # Establece colores para usar en la ventana principal
        light_bg_color = "#ffffff"  # Background color for light theme
        dark_bg_color = "#313131"  # Background color for dark theme

        # Parsea los argumentos para obtener el tema
        parser = argparse.ArgumentParser()
        parser.add_argument("--theme", help="Tema: 'dark' o 'light'")
        args = parser.parse_args()

        # Configura el tema de acuerdo al argumento pasado
        if args.theme == "dark":
            # Configurar tema oscuro
            style.theme_use("forest-dark")
        elif args.theme == "light":
            # Configurar tema claro
            style.theme_use("forest-light")
        else:
            # Configurar tema oscuro
            style.theme_use("forest-dark")

        # Establecer el icono personalizado

        icon_dir = os.path.join(get_executable_directory(), "Assets/Icons")
        icon_file = "file_icon.ico"

        icono_path = os.path.join(icon_dir, icon_file)
        root.iconbitmap(icono_path)

        # Creación de un frame principal
        frame = ttk.Frame(root)
        frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)  # Rellenar y expandir el marco principal

        # Frame para los botones de carga de carpetas
        folder_buttons_frame = ttk.Frame(frame)
        folder_buttons_frame.grid(row=0, column=0, sticky="news")
        folder_buttons_frame.columnconfigure(2, weight=1)  # La columna 2 (donde está el Checkbutton) se expandirá

        # Botones para cargar las carpetas
        self.left_button = ttk.Button(folder_buttons_frame, text="Cargar primera carpeta", command=self.load_left_dir)
        self.left_button.grid(row=0, column=0, padx=5, pady=5, sticky="news")

        self.right_button = ttk.Button(folder_buttons_frame, text="Cargar segunda carpeta", command=self.load_right_dir)
        self.right_button.grid(row=0, column=1, padx=5, pady=5, sticky="news")

        # Switch para el cambio de modo light or dark
        mode_switch = ttk.Checkbutton(
            folder_buttons_frame, text="Mode", style="Switch", command=toggle_mode) # style="Switch" le da el estilo
        mode_switch.grid(row=0, column=2, padx=(0,20), pady=10, sticky="e")

        # Si el parámetro theme es no es nulo (= se ejecuta desde la botonera), oculta el mode_switch
        if args.theme  is not None:
            mode_switch.grid_remove()
            root.configure(bg=dark_bg_color) if args.theme == 'dark' else root.configure(bg=light_bg_color)

        # Frame para los treeviews de los archivos en las carpetas
        treeviews_frame = ttk.Frame(frame)
        treeviews_frame.grid(row=1, column=0, padx=5, pady=(0, 5), sticky="news")

        # Treeview para la primera carpeta
        self.left_tree = ttk.Treeview(treeviews_frame, columns=("Files"), show='headings')
        self.left_tree.heading("Files", text="Archivos en la primera carpeta")
        self.left_tree.grid(row=0, column=0, padx=5, pady=5, sticky=(tk.W, tk.E, tk.N, tk.S))
        self.left_tree.column("Files", width=340, minwidth=340, stretch=False)
        self.left_scrollbar = ttk.Scrollbar(treeviews_frame, orient="vertical", command=self.left_tree.yview)
        self.left_scrollbar.grid(row=0, column=1, padx=(0,5), pady=5, sticky=(tk.N, tk.S))
        self.left_tree.configure(yscrollcommand=self.left_scrollbar.set)

        # Treeview para la segunda carpeta
        self.right_tree = ttk.Treeview(treeviews_frame, columns=("Files"), show='headings')
        self.right_tree.heading("Files", text="Archivos en la segunda carpeta")
        self.right_tree.grid(row=0, column=2, padx=5, pady=5, sticky=(tk.W, tk.E, tk.N, tk.S))
        self.right_tree.column("Files", width=340, minwidth=340, stretch=False)
        self.right_scrollbar = ttk.Scrollbar(treeviews_frame, orient="vertical", command=self.right_tree.yview)
        self.right_scrollbar.grid(row=0, column=3, padx=(0,5), pady=5, sticky=(tk.N, tk.S))
        self.right_tree.configure(yscrollcommand=self.right_scrollbar.set)

        # Frame para el botón de conciliación
        reconcile_button_frame = ttk.Frame(frame)
        reconcile_button_frame.grid(row=2, column=0, padx=5, pady=(0, 5), sticky=(tk.W, tk.E))

        # Botón de conciliación
        self.reconcile_button = ttk.Button(reconcile_button_frame, text="Conciliar", command=self.compare_files)
        self.reconcile_button.pack(padx=5, pady=5, fill=tk.X)

        # Frame para los treeviews resultantes de la conciliación
        result_treeviews_frame = ttk.Frame(frame)
        result_treeviews_frame.grid(row=3, column=0, padx=5, pady=(0, 5), sticky="news")

        # Treeview para los archivos solo en la primera carpeta
        self.left_only_tree = ttk.Treeview(result_treeviews_frame, columns=("Files"), show='headings')
        self.left_only_tree.heading("Files", text="Archivos solo en la primera carpeta")
        self.left_only_tree.grid(row=0, column=0, padx=5, pady=5, sticky=(tk.W, tk.E, tk.N, tk.S))
        self.left_only_tree.column("Files", width=325, minwidth=325, stretch=False)
        self.left_only_scrollbar = ttk.Scrollbar(result_treeviews_frame, orient="vertical", command=self.left_only_tree.yview)
        self.left_only_scrollbar.grid(row=0, column=1, padx=(0,5), pady=5, sticky=(tk.N, tk.S))
        self.left_only_tree.configure(yscrollcommand=self.left_only_scrollbar.set)

        # Treeview para los archivos solo en la segunda carpeta
        self.right_only_tree = ttk.Treeview(result_treeviews_frame, columns=("Files"), show='headings')
        self.right_only_tree.heading("Files", text="Archivos solo en la segunda carpeta")
        self.right_only_tree.grid(row=0, column=2, padx=5, pady=5, sticky=(tk.W, tk.E, tk.N, tk.S))
        self.right_only_tree.column("Files", width=325, minwidth=325, stretch=False)
        self.right_only_scrollbar = ttk.Scrollbar(result_treeviews_frame, orient="vertical", command=self.right_only_tree.yview)
        self.right_only_scrollbar.grid(row=0, column=3, padx=(0,5), pady=5, sticky=(tk.N, tk.S))
        self.right_only_tree.configure(yscrollcommand=self.right_only_scrollbar.set)

        # Botón para exportar a Excel
        export_button = ttk.Button(frame, text="Exportar a Excel", command=self.export_to_excel)
        export_button.grid(row=4, column=0, padx=10, pady=10)

        # Configurar el redimensionamiento de la cuadrícula
        frame.columnconfigure(0, weight=1)
        frame.rowconfigure(0, weight=0)
        frame.rowconfigure(1, weight=1)
        frame.rowconfigure(2, weight=0)
        frame.rowconfigure(3, weight=1)

        # Configurar el redimensionamiento de la cuadrícula para treeviews_frame
        treeviews_frame.columnconfigure(0, weight=1)
        treeviews_frame.columnconfigure(2, weight=1)
        treeviews_frame.rowconfigure(0, weight=1)

        # Configurar el redimensionamiento de la cuadrícula para result_treeviews_frame
        result_treeviews_frame.columnconfigure(0, weight=1)
        result_treeviews_frame.columnconfigure(2, weight=1)
        result_treeviews_frame.rowconfigure(0, weight=1)

    def load_left_dir(self):
        self.left_dir = filedialog.askdirectory(title="Seleccionar la primera carpeta")
        self.left_files = self.file_list(self.left_dir, self.left_dir)
        self.populate_treeview(self.left_tree, self.left_files)

    def load_right_dir(self):
        self.right_dir = filedialog.askdirectory(title="Seleccionar la segunda carpeta")
        self.right_files = self.file_list(self.right_dir, self.right_dir)
        self.populate_treeview(self.right_tree, self.right_files)

    def populate_treeview(self, tree, files):
        for row in tree.get_children():
            tree.delete(row)
        for file in files:
            tree.insert("", "end", values=(file,))

    def file_list(self, dir_path, root_dir):
        archivos = []
        for arch in scandir(dir_path):
            if arch.is_file():
                archivos.append("\\" + relpath(abspath(arch.path), start=root_dir).replace("/", "\\"))
            elif arch.is_dir():
                archivos.extend(self.file_list(join(dir_path, arch.path), root_dir))
        return archivos

    def compare_files(self):
        only_in_left = [file for file in self.left_files if file not in self.right_files]
        only_in_right = [file for file in self.right_files if file not in self.left_files]

        # Llenar Treeview de archivos solo en la primera carpeta
        self.populate_treeview(self.left_only_tree, only_in_left)

        # Llenar Treeview de archivos solo en la segunda carpeta
        self.populate_treeview(self.right_only_tree, only_in_right)


    def export_to_excel(self):
        only_in_left = [file for file in self.left_files if file not in self.right_files]
        only_in_right = [file for file in self.right_files if file not in self.left_files]

        # Crear un nuevo libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"

        # Configurar el nombre de las columnas
        ws.cell(row=1, column=1, value="First folder only")
        ws.cell(row=1, column=2, value="Second folder only")

        # Escribir los datos en el libro de Excel
        for i, file in enumerate(only_in_left, start=2):
            ws.cell(row=i, column=1, value=file)
        for i, file in enumerate(only_in_right, start=2):
            ws.cell(row=i, column=2, value=file)

        # Guardar el archivo Excel
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")], title="Guardar archivo de Excel")
        if file_path:
            wb.save(file_path)

        # Abrir el archivo de Excel con la aplicación predeterminada
        # Encerrar el nombre del archivo entre comillas para manejar nombres con espacios
        os.system('start excel.exe "' + file_path + '"')

if __name__ == "__main__":
    root = ThemedTk(theme="black")
    
   

    app = FileReconcilerApp(root)
    root.geometry("800x600")
    root.resizable(False, False)  # Deshabilitar el redimensionamiento
    root.mainloop()
